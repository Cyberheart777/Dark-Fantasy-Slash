#!/usr/bin/env python3
"""Generate an Excel spreadsheet of all power-ups from UpgradeData.ts"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

upgrades = [
    # (Category, ID, Name, Icon, Description, Rarity, MaxStacks, Classes)
    # ── UNIVERSAL ──
    ("Universal", "damage_boost", "Whetstone", "⚔️", "+15% weapon damage", "Common", 8, "All"),
    ("Universal", "attack_speed_boost", "Bladestorm", "🌪️", "+12% attack speed", "Common", 6, "All"),
    ("Universal", "max_health_boost", "Iron Constitution", "❤️", "+20 max health and heal 20 HP", "Common", 10, "All"),
    ("Universal", "crit_chance_boost", "Critical Eye", "🎯", "+6% critical strike chance", "Common", 6, "All"),
    ("Universal", "move_speed_boost", "Windwalker", "💨", "+12% movement speed", "Common", 5, "All"),
    ("Universal", "lifesteal_start", "Blood Price", "🩸", "Gain 4% lifesteal on hits", "Rare", 1, "All"),
    ("Universal", "lifesteal_boost", "Bloodlord", "🩸", "+2% lifesteal", "Common", 5, "All"),
    ("Universal", "double_strike", "Twin Fang", "⚡", "+18% chance to strike twice", "Rare", 4, "All"),
    ("Universal", "armor_boost", "Tempered Plate", "🛡️", "+5 armor (reduces incoming damage)", "Common", 6, "All"),
    ("Universal", "health_regen", "Troll's Blood", "✨", "+1.0 HP regen per second", "Common", 5, "All"),
    ("Universal", "dash_cooldown", "Phantom Step", "🌀", "-20% dash cooldown", "Common", 5, "All"),
    ("Universal", "xp_gain_boost", "Scholar's Insight", "📖", "+25% XP from all sources", "Common", 4, "All"),
    ("Universal", "berserker_rage", "Glass Cannon", "💥", "+20% damage but -10% max health", "Rare", 3, "All"),
    ("Universal", "iron_skin", "Iron Skin", "🪬", "+5% dodge chance", "Common", 4, "All"),
    ("Universal", "attack_range_boost", "Executioner's Reach", "🗡️", "+1 attack range", "Common", 4, "All"),
    ("Universal", "soul_feast", "Soul Feast", "👻", "Heal 8 HP on every kill", "Common", 5, "All"),
    ("Universal", "wraithplate", "Wraithplate", "🦴", "+10 armor", "Common", 4, "All"),
    ("Universal", "overclock", "Overclock", "⚡", "+5% attack speed and +5% move speed", "Common", 5, "All"),

    # ── WARRIOR ──
    ("Warrior", "cleave_start", "Wide Swing", "🪓", "Attacks now cleave — wider arc (+30°)", "Rare", 1, "Warrior"),
    ("Warrior", "cleave_boost", "Arc Master", "🪓", "+20° attack arc", "Common", 5, "Warrior"),
    ("Warrior", "blood_momentum", "Blood Momentum", "🔴", "Each consecutive hit increases damage by +3% (max 60%). Resets after 2s.", "Epic", 1, "Warrior"),
    ("Warrior", "earthbreaker", "Earthbreaker", "🌋", "Every 5th hit slams the ground — AoE damage to all nearby enemies.", "Epic", 1, "Warrior"),
    ("Warrior", "iron_reprisal", "Iron Reprisal", "💢", "Taking damage releases a shockwave dealing 15% of your max HP as damage.", "Rare", 1, "Warrior"),
    ("Warrior", "fortress", "Fortress", "🏰", "Gain +2 armor/sec while standing still (max +20).", "Rare", 1, "Warrior"),
    ("Warrior", "war_cry", "Battle Roar", "📯", "War Cry damage bonus increased to +35% for 6 seconds.", "Rare", 1, "Warrior"),
    ("Warrior", "bloodforge", "Bloodforge", "🩸", "Each kill grants +1 max HP (capped at +20).", "Rare", 1, "Warrior"),
    ("Warrior", "weakening_blows", "Weakening Blows", "💀", "Melee hits reduce enemy damage by 2%.", "Rare", 3, "Warrior"),
    ("Warrior", "serrated_edge", "Serrated Edge", "🩸", "Critical hits apply a bleed: 6 damage/sec for 3s.", "Rare", 3, "Warrior"),
    ("Warrior", "concussive_charge", "Concussive Charge", "💥", "Dash knockback distance +50%. Knocked enemies take damage.", "Rare", 2, "Warrior"),

    # ── MAGE ──
    ("Mage", "chain_lightning", "Chain Lightning", "⚡", "Projectile hits bounce to 2 nearby enemies for 60% damage.", "Rare", 3, "Mage"),
    ("Mage", "spell_echo", "Spell Echo", "🔮", "+25% chance to double-cast your projectile.", "Rare", 3, "Mage"),
    ("Mage", "arcane_fracture", "Arcane Fracture", "💎", "Enemies killed by your projectiles explode into 3 mini-projectiles.", "Epic", 1, "Mage"),
    ("Mage", "mana_shield", "Mana Shield", "🛡️", "Absorb 25% of incoming damage as a magic barrier.", "Rare", 1, "Mage"),
    ("Mage", "singularity", "Singularity", "🌀", "Every 14s create a vortex pulling enemies inward for 3s.", "Epic", 1, "Mage"),
    ("Mage", "frost_armor", "Frost Armor", "❄️", "Enemies that hit you are slowed by 35% for 2s.", "Rare", 1, "Mage"),
    ("Mage", "arcane_detonation", "Arcane Detonation", "💥", "Orbs explode on expiry for 60% AoE damage.", "Epic", 1, "Mage"),
    ("Mage", "gravity_orbs", "Gravity Orbs", "🌀", "Orbs pull nearby enemies toward their flight path.", "Epic", 1, "Mage"),
    ("Mage", "overcharged_orbs", "Overcharged Orbs", "⚡", "Orbs gain up to +80% damage at max range.", "Rare", 1, "Mage"),
    ("Mage", "residual_field", "Residual Field", "🔮", "Orbs leave a damaging trail that burns enemies.", "Epic", 1, "Mage"),
    ("Mage", "extra_orb", "Arcane Barrage", "🟣", "Fire +1 additional orb per attack.", "Rare", 3, "Mage"),
    ("Mage", "volatile_blink", "Volatile Blink", "💥", "Blink afterimage now explodes for 1× damage in a wide radius.", "Epic", 1, "Mage"),
    ("Mage", "projectile_size", "Amplified Orbs", "🔵", "+20% projectile collision radius.", "Common", 3, "Mage"),
    ("Mage", "split_bolt", "Split Bolt", "🔀", "+1 orb per attack but -25% damage. Trades focus for spread.", "Rare", 1, "Mage"),

    # ── ROGUE ──
    ("Rogue", "shadow_step", "Shadow Step", "👤", "Dash cooldown resets on kill.", "Epic", 1, "Rogue"),
    ("Rogue", "venom_stack", "Venom Stack", "🐍", "Attacks apply poison: 4 damage/sec per stack. Spreads on death.", "Rare", 3, "Rogue"),
    ("Rogue", "crit_cascade", "Crit Cascade", "💫", "Critical hits boost crit chance by +12% for 3s.", "Epic", 1, "Rogue"),
    ("Rogue", "phantom_blades", "Phantom Blades", "🗡️", "Each attack fires 2 extra spectral daggers at wide angles.", "Epic", 1, "Rogue"),
    ("Rogue", "evasion_matrix", "Evasion Matrix", "🌫️", "Successful dodge grants 1s invisibility + guaranteed crit.", "Rare", 1, "Rogue"),
    ("Rogue", "blade_orbit", "Blade Orbit", "🔄", "3 daggers spin around you, damaging nearby enemies.", "Rare", 2, "Rogue"),
    ("Rogue", "extra_daggers", "Fan of Knives", "🔪", "Fire +1 additional dagger per attack.", "Rare", 3, "Rogue"),
    ("Rogue", "toxic_dash", "Toxic Dash", "☠️", "Dash applies 3 poison stacks instead of 1.", "Rare", 1, "Rogue"),
    ("Rogue", "deep_wounds", "Deep Wounds", "🧪", "Poison damage and duration increased by 50%.", "Rare", 2, "Rogue"),

    # ── RELICS ──
    ("Relic", "relic_soulfire", "Soulfire Blade", "🔥", "Kills have a 20% chance to explode, dealing 1.5× your damage nearby.", "Epic", 1, "All"),
    ("Relic", "relic_vampiric", "Vampiric Shroud", "🧛", "+4% lifesteal. Heal 2 HP on kill.", "Epic", 1, "All"),
    ("Relic", "relic_phantom_echo", "Phantom Echo", "👁️", "Every 5th attack fires a free ghost strike for 50% damage.", "Epic", 1, "Warrior"),
    ("Relic", "relic_deaths_bargain", "Death's Bargain", "💀", "Once per run, survive a lethal blow with 1 HP. 1.5s invincibility.", "Epic", 1, "All"),
    ("Relic", "relic_abyss_crown", "Abyss Crown", "👑", "+40% XP gain. Cursed: you take 20% more damage.", "Epic", 1, "All"),
    ("Relic", "relic_blood_covenant", "Blood Covenant", "🩸", "Sacrifice 20% max HP. Deal +25% more damage permanently.", "Epic", 1, "All"),
    ("Relic", "relic_storm_heart", "Storm Heart", "⚡", "Every 18s, lightning strikes up to 8 enemies for 1.2× your damage.", "Epic", 1, "All"),
    ("Relic", "relic_iron_oath", "Iron Oath", "⚙️", "+25 armor and +20% max HP. Your dash is disabled.", "Epic", 1, "Warrior"),
]

wb = Workbook()
ws = wb.active
ws.title = "Power-Ups"

# ── Colors ──
HEADER_FILL = PatternFill(start_color="2D2D2D", end_color="2D2D2D", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=12)

CAT_FILLS = {
    "Universal": PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid"),
    "Warrior":   PatternFill(start_color="2E1A1A", end_color="2E1A1A", fill_type="solid"),
    "Mage":      PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid"),
    "Rogue":     PatternFill(start_color="1A2E1A", end_color="1A2E1A", fill_type="solid"),
    "Relic":     PatternFill(start_color="2E2A1A", end_color="2E2A1A", fill_type="solid"),
}
CAT_FONTS = {
    "Universal": Font(name="Calibri", color="AAAAFF", size=11),
    "Warrior":   Font(name="Calibri", color="FF9999", size=11),
    "Mage":      Font(name="Calibri", color="CC99FF", size=11),
    "Rogue":     Font(name="Calibri", color="99FF99", size=11),
    "Relic":     Font(name="Calibri", color="FFD700", size=11),
}

RARITY_FONTS = {
    "Common": Font(name="Calibri", color="AAAAAA", size=11),
    "Rare":   Font(name="Calibri", color="6699FF", size=11, bold=True),
    "Epic":   Font(name="Calibri", color="CC66FF", size=11, bold=True),
}

DARK_BG = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid")
DEFAULT_FONT = Font(name="Calibri", color="DDDDDD", size=11)
NAME_FONT = Font(name="Calibri", color="FFFFFF", size=11, bold=True)

thin_border = Border(
    left=Side(style="thin", color="444444"),
    right=Side(style="thin", color="444444"),
    top=Side(style="thin", color="444444"),
    bottom=Side(style="thin", color="444444"),
)

# ── Headers ──
headers = ["Category", "Icon", "Name", "Description", "Rarity", "Max Stacks", "Classes"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

# ── Data rows ──
for row_idx, (cat, uid, name, icon, desc, rarity, stacks, classes) in enumerate(upgrades, 2):
    data = [cat, icon, name, desc, rarity, stacks, classes]
    for col, val in enumerate(data, 1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.fill = DARK_BG
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=(col == 4))

        if col == 1:  # Category
            cell.font = CAT_FONTS.get(cat, DEFAULT_FONT)
            cell.fill = CAT_FILLS.get(cat, DARK_BG)
        elif col == 3:  # Name
            cell.font = NAME_FONT
        elif col == 5:  # Rarity
            cell.font = RARITY_FONTS.get(rarity, DEFAULT_FONT)
        elif col == 6:  # Max Stacks
            cell.font = DEFAULT_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.font = DEFAULT_FONT

# ── Column widths ──
ws.column_dimensions["A"].width = 12   # Category
ws.column_dimensions["B"].width = 5    # Icon
ws.column_dimensions["C"].width = 22   # Name
ws.column_dimensions["D"].width = 62   # Description
ws.column_dimensions["E"].width = 10   # Rarity
ws.column_dimensions["F"].width = 12   # Max Stacks
ws.column_dimensions["G"].width = 10   # Classes

# ── Freeze header row ──
ws.freeze_panes = "A2"

# ── Auto-filter ──
ws.auto_filter.ref = f"A1:G{len(upgrades) + 1}"

out = "/home/user/Dark-Fantasy-Slash/Dungeon_Requiem_Power_Ups.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Total upgrades: {len(upgrades)}")
